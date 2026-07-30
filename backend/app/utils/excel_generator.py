"""
VyaparPro — Excel Report Generator
Generates .xlsx workbooks for accounting/business reports using openpyxl.
Mirrors the (summary, tables) shape used by generate_report_pdf so both
formats are built from the exact same data — no duplicated report logic.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, color="666666")
SECTION_FONT = Font(bold=True, size=11, color="0066CC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_TOTAL_PREFIXES = ("total", "closing", "net", "grand total", "gross profit", "operating profit", "net profit")


def generate_report_xlsx(
    title: str,
    subtitle: str,
    summary: list,
    tables: list,
    company: dict | None = None,
    period: str = "",
) -> bytes:
    """
    summary: list of (label, value) OR (label, value, accent) — accent is ignored in Excel.
    tables:  list of (section_title, headers, rows) — rows is list[list[str]].
    Each table is written to its own worksheet (keeps wide reports readable);
    a single-table report just gets one sheet.
    """
    wb = Workbook()
    wb.remove(wb.active)

    company_name = (company or {}).get("legal_name", "")

    for section_title, headers, rows in tables:
        sheet_name = (section_title or "Report")[:31].replace("/", "-")
        # Excel disallows duplicate sheet names — disambiguate if needed.
        base_name, i = sheet_name, 1
        while sheet_name in wb.sheetnames:
            i += 1
            sheet_name = f"{base_name[:28]} {i}"
        ws = wb.create_sheet(sheet_name)

        row = 1
        if company_name:
            ws.cell(row=row, column=1, value=company_name).font = TITLE_FONT
            row += 1
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        if subtitle or period:
            ws.cell(row=row, column=1, value=subtitle or period).font = SUBTITLE_FONT
            row += 1
        row += 1

        # Summary KPIs only on the first sheet
        if summary and section_title == tables[0][0]:
            for item in summary:
                label, value = item[0], item[1]
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
            row += 1

        ws.cell(row=row, column=1, value=section_title).font = SECTION_FONT
        row += 1

        if not rows:
            ws.cell(row=row, column=1, value="No data for this period")
            continue

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        header_row = row
        row += 1

        for r in rows:
            for col, val in enumerate(r, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = BORDER
                if col > 1:
                    c.alignment = Alignment(horizontal="right")
            first_val = str(r[0]).strip().lower() if r else ""
            if first_val.startswith(_TOTAL_PREFIXES):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).font = TOTAL_FONT
            row += 1

        for col in range(1, len(headers) + 1):
            max_len = max(
                [len(str(headers[col - 1]))]
                + [len(str(r[col - 1])) for r in rows if len(r) >= col]
                + [0]
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, 12), 40)

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()